from collections import defaultdict
from openpyxl import load_workbook
from config import arquivo_excel, aba_entradas, aba_gastos


def gerar_resumo_mensal():
    wb = load_workbook(arquivo_excel)
    ws_entradas = wb[aba_entradas]
    ws_gastos = wb[aba_gastos]
    ws_resumo = wb['RESUMO']

    resumo = defaultdict(lambda: {'entradas': 0, 'gastos': 0})

    for row in ws_entradas.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            continue
        dia, mes, ano, *_, valor, _ = row

        if mes is None or ano is None or valor is None:
            continue
        resumo[(int(mes), int(ano))]['entradas'] += float(valor)

    for row in ws_gastos.iter_rows(min_row=2, values_only=True):
        if len(row) < 10:
            continue

        (
            dia,
            mes,
            ano,
            descricao,
            categoria,
            valor_total,
            forma_pagamento,
            parcelado,
            numero_parcelas,
            valor_parcela
        ) = row

        if mes is None or ano is None:
            continue

        if parcelado:
            if valor_parcela is None:
                continue
            resumo[(int(mes), int(ano))]['gastos'] += float(valor_parcela)
        
        else:
            if valor_total is None:
                continue
            resumo[(int(mes), int(ano))]['gastos'] += float(valor_total)

    if ws_resumo.max_row > 1:
        ws_resumo.delete_rows(2, ws_resumo.max_row)

    for (mes, ano) in sorted(resumo.keys(), key=lambda x: (x[1], x[0])):
        entradas = round(resumo[(mes, ano)]['entradas'], 2)
        gastos = round(resumo[(mes, ano)]['gastos'], 2)
        saldo = round(entradas - gastos, 2)

        ws_resumo.append([mes, ano, entradas, gastos, saldo])
    
    wb.save(arquivo_excel)
